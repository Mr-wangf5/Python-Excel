from tkinter import *
import tkinter as tk
import openpyxl    # 操作excel表格
import pymysql   # 链接数据库
import tkinter.font as tkFont    # 字体
from tkinter import ttk
from tkinter.filedialog import askopenfilename, askdirectory    # 系统文件操作
from tkinter import messagebox
from pandas import DataFrame    # 写入数据表
from commodity import commodity     # 商品实体类


# 创建主页ui

class Application(Frame):
    # 初始化
    def __init__(self, master=None):
        self.master = master
        self.pack()
        self.createWidget()
        # 链接数据库
        self.conn = pymysql.connect(host='localhost', user='root', password='123456', database='db_shopping')
        self.cursor = self.conn.cursor()

    def createWidget(self):
        # 左侧背景图片
        global photo_bg
        photo_bg = PhotoImage(file='./images/login.gif')
        label_bg = Label(self, image=photo_bg)
        label_bg.pack()
        self.frame_left = Frame(self, bg='#68B8BE', width=200, height=540).place(x=0, y=0)

        # logo图
        global photo_Avatar
        photo_Avatar = PhotoImage(file='./images/login.gif')
        label_Avatar = Label(self.frame_left, bg='#68B8BE', image=photo_Avatar)
        label_Avatar.place(x=23, y=25)

        # 功能按钮
        menu_font= tkFont.Font(family='宋体', size=10)
        Button(self.frame_left, text='商品信息', font=menu_font, fg='#000000', bg='#ffffff',
               command=self.index_info).place(x=30, y=193, width=140, height=30)
        Button(self.frame_left, text='查询商品', font=menu_font, fg='#000000', bg='#ffffff',
               command=self.index_inquire).place(x=30, y=243, width=140, height=30)
        Button(self.frame_left, text='添加商品', font=menu_font, fg='#000000', bg='#ffffff',
               command=self.index_add).place(x=30, y=293, width=140, height=30)
        Button(self.frame_left, text='修改商品', font=menu_font, fg='#000000', bg='#ffffff',
               command=self.index_modify).place(x=30, y=343, width=140, height=30)
        Button(self.frame_left, text='删除商品', font=menu_font, fg='#000000', bg='#ffffff',
               command=self.index_delete).place(x=30, y=393, width=140, height=30)
        Button(self.frame_left, text='关于', font=menu_font, fg='#000000', bg='#ffffff',
               command=self.index_about).place(x=30, y=443, width=140, height=30)
        Button(self.frame_left, text='退出', font=menu_font, fg='#000000', bg='#ffffff',
               command=self.index_quit).place(x=30, y=493, width=140, height=30)

        # 右侧背景图
        self.frame_right = Frame(self, width=600, height=540).place(x=200, y=0)
        global photo_index
        photo_index = PhotoImage(file='./images/bg.gif')
        Label(self.frame_right, image=photo_index).place(x=200, y=0)

    # 01 查看商品信息
    def index_info(self):
        info = tk.Toplevel()
        w = 600
        h = 400
        x = (info.winfo_screenwidth() - w) /2
        y = (info.winfo_screenheight() - h)/2
        info.geometry('%dx%d+%d+%d' % (w, h, x, y))
        info.title('商品信息')
        info.resizable(width=False, height=False)
        info_frame = Frame(info, bg='#68B8BE', width=600, height=55)
        info_frame.place(x=0, y=0)
        inquire_font_1 = tkFont.Font(family='宋体', size=25, weight=tkFont.BOLD)
        Label(info_frame, text='商品信息', font=inquire_font_1, bg='#68B8BE').place(x=230, y=10)
        entries = []
        sql = "select * from commodity"
        self.cursor.execute(sql)
        results = self.cursor.fetchall()
        com_list =[]
        for result in results:
            com_li = commodity(result[0], result[1], result[2], result[3])
            com_list.append(com_li)
        for i in com_list:
            args = (i.get_commodity()['com_code'], i.get_commodity()['com_name'],
                    i.get_commodity()['com_price'], i.get_commodity()['com_stock'])
            entries.append(args)

        tabel_frame = Frame(info)
        tabel_frame.place(x=0, y=55, width=600, height=345)
        yscroll = Scrollbar(tabel_frame, orient=VERTICAL)
        style = ttk.Style()
        style.configure('Treeview.Heading', font=(None, 12))
        tree = ttk.Treeview(
            master=tabel_frame,
            columns=('商品编号', '商品名称', '商品价格', '商品库存'),
            yscrollcommand=yscroll.set
        )
        yscroll.configure(command=tree.yview)
        yscroll.pack(side=RIGHT, fill=Y)
        tree['show'] = 'headings'
        tree.heading('#1', text='商品编号')
        tree.heading('#1', text='商品名称')
        tree.heading('#1', text='商品价格')
        tree.heading('#1', text='商品库存')
        tree.column('#1', stretch=YES, width=150, minwidth=150, anchor='center')
        tree.column('#2', stretch=YES, width=150, minwidth=150, anchor='center')
        tree.column('#3', stretch=YES, width=150, minwidth=150, anchor='center')
        tree.column('#4', stretch=YES, width=150, minwidth=150, anchor='center')
        tree.pack(fill=BOTH, expand=1)
        for entry in entries:
            tree.insert('', 'end', values=(entry[0], entry[1], entry[2], entry[3]))
        info.mainloop()

    # 查询商品
    def index_inquire(self):
        inquire = tk.Toplevel()
        w = 600
        h = 400
        x = (inquire.winfo_screenwidth() -w) /2
        y = (inquire.winfo_screenheight() - h )/2
        global com_name
        global com_price
        global com_stock
        global com_name_label
        global com_price_label
        global com_stock_label
        com_name = '查询中'
        com_price = '查询中'
        com_stock = '查询中'
        inquire.geometry('%dx%d+%d+%d' % (w, h, x, y))
        inquire.title('查询商品')
        inquire.resizable(width=False, height=False)
        inquire_frame = Frame(inquire, bg='#68B8BE', width=600, height=400)
        inquire_frame.pack()
        inquire_font_1 = tkFont.Font(family='宋体', size=15, weight=tkFont.BOLD)
        Label(inquire_frame, text='查询商品信息', font=inquire_font_1, bg='#68B8BE').place(x=200, y=150)
        inquire_font_2 = tkFont.Font(family='宋体', size=15)
        Label(inquire_frame, text='商品编号', font=inquire_font_2, bg='#68B8BE').place(x=200, y=90)
        Label(inquire_frame, text='商品名称', font=inquire_font_2, bg='#68B8BE').place(x=200, y=140)
        Label(inquire_frame, text='商品价格', font=inquire_font_2, bg='#68B8BE').place(x=200, y=290)
        Label(inquire_frame, text='商品库存', font=inquire_font_2, bg='#68B8BE').place(x=200, y=340)
        com_code = tk.Scrollbar()
        Entry(inquire_frame, highlightthickness=1, font=('宋体', 15), bg='#F3F3F4',
              textvariable=com_code).place(x=300, y=90, width=200, height=30)
        com_name_label = Label(inquire_frame, text=com_name, font=inquire_font_2, bg='#68B8BE')
        com_name_label.place(x=300, y=150)
        com_price_label = Label(inquire_frame, text=com_price, font=inquire_font_2, bg='#68B8BE')
        com_price_label.place(x=300, y=210)
        com_stock_label = Label(inquire_frame, text=com_stock, font=inquire_font_2, bg='#68B8BE')
        com_stock_label.place(x=300, y=150)

        Button(inquire_frame, text='立即查询', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command = lambda :self.inquire_com(com_code)).place(x=180, y=325, height=40)
        inquire.mainloop()

    # 02-1 数据库中查询商品，返回结果
    def inquire(self, com_code):
        sql = "select * fom commodity where com_code=%s" % com_code.get()
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.conn.commit()
        if len(result) == 0 :
            com_name_label.config(text='未查询到商品')
            com_price_label.config(text='未查询到商品')

        else:
            com = commodity(result[0][0], result[0][1], result[0][2], result[0][3])
            if com is False:
                com_code.set('未查询到商品')
                com_name_label.config(text='未查询到商品')
                com_price_label.config(text='未查询到商品')
                com_stock_label.config(text='未查询到商品')
            else:
                com_name_label.config(text=com.get_commodity()['com_name'])
                com_price_label.config(text=com.get_commodity()['com_price'])
                com_stock_label.config(text=com.get_commodity()['com_stock'])

    # 03 添加商品界面
    def index_add(self):
        add = Toplevel()
        w = 600
        h = 400
        x = (add.winfo_screenwidth() -w) /2
        y = (add.winfo_screenheight() - h )/2
        self.com_code = tk.StringVar()
        self.com_name = tk.StringVar()
        self.com_price = tk.StringVar()
        self.com_stock = tk.StringVar()
        add.geometry('%dx%d+%d+%d' % (w, h, x, y))
        add.title('修改商品')
        add.resizable(width=False, height=False)
        add_frame = Frame(add, bg='#68B8BE', width=600, height=400)
        add_frame.pack()
        add_font_1 = tkFont.Font(family='宋体', size=15, weight=tkFont.BOLD)
        Label(add_frame, text='添加商品信息', font=add_font_1, bg='#68B8BE').place(x=200, y=150)
        add_font_2 = tkFont.Font(family='宋体', size=15)
        modify_font_3 = tkFont.Font(family='宋体', size=9)
        Label(add_frame, text='注意：商品编号不可重复', font=self.com_code).place(x=235, y=125)
        Label(add_frame, text='商品名称', font=add_font_2, bg='#68B8BE').place(x=100, y=150)
        Label(add_frame, text='商品价格', font=add_font_2, bg='#68B8BE').place(x=100, y=210)
        Label(add_frame, text='商品库存', font=add_font_2, bg='#68B8BE').place(x=100, y=270)
        Entry(add_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_code).\
            place(x=300, y=90, height=30, width=200)
        Entry(add_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_name). \
            place(x=300, y=150, height=30, width=200)
        Entry(add_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_price). \
            place(x=300, y=210, height=30, width=200)
        Entry(add_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_stock). \
            place(x=300, y=270, height=30, width=200)
        Button(add_frame, text='批量添加', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command=self.adds_com_up).place(x=70, y=325, width=120, height=40)
        Button(add_frame, text='立即添加', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command=self.add_com).place(x=230, y=325, width=120, height=40)
        Button(add_frame, text='下载模板', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command=self.adds_com_download).place(x=390, y=325, width=120, height=40)

    # 03-1 数据库中添加商品（上传文件方式）
    def adds_com_up(self):
        path = askopenfilename(title='上传文件', initialdir='D', filetypes=[('Excel表格', '.xlsx')])
        file = openpyxl.load_workbook(path)
        ws = file[file.sheetnames[0]]       # 第一个表
        excel_max_row = ws.max_row     # 最大行
        list_com = []
        for row in range(2, excel_max_row+1):
            com_code = ws.cell(row, 1).value
            com_name = ws.cell(row, 2).value
            com_price = ws.cell(row, 1).value
            com_stock = ws.cell(row, 2).value
            com = commodity(com_code, com_name, com_price, com_stock)
            com_dict = com.get_commodity()
            list_com.append(com_dict)
        results = []
        for com in list_com:
            com_code = com['com_code'][0]
            com_name = com['com_name'][0]
            com_price = com['com_price'][0]
            com_stock = com['com_stock'][0]
            sql = 'insert into commodity(com_code, com_name, com_price, com_stock) values (%s, %s, %s, %s)' % (
                com_code, com_name, com_price, com_stock)
            self.cursor.execute(sql)
            result = self.cursor.fetchall()
            results.append(result)

        self.conn.commit()
        if len(results) == excel_max_row -1:
            messagebox.showinfo('提示', '商品批量导入成功！')
        elif len(results) !=0 and len(results) != excel_max_row-1:
            messagebox.showinfo('提示', '导入成功%s条商品数据， 共%s条商品数据！'% (len(results), excel_max_row-1))
        else:
            messagebox.showinfo('提示', '商品批量导入失败')

    # 03-2 数据库添加商品（直接输入的方式）
    def add_com(self):
        sql = 'insert into commodity(com_code, com_name, com_price, com_stock) values (%s, %s, %s, %s)' % (
            self.com_code.get(), self.com_name.get(), self.com_price.get(), self.com_stock.get())
        try:
            self.cursor.execute(sql)
            self.conn.commit()
            messagebox.showinfo('提示', '添加商品成功')
        except:
            messagebox.showinfo('提示', '添加商品失败')

    # 03-3 数据库中添加数据（下载模板）
    def add_com_download(self):
        path = askdirectory(title='下载到')
        com = commodity('使用前请删除本行', 'tip:商品编号不允许重复', '1', '1')
        dic_com = com.get_commodity()
        list_com = [dic_com]
        com_code = []
        com_name = []
        com_price = []
        com_stock =[]
        for li_com in list_com:
            com_code.append(li_com['com_code'])
            com_name.append(li_com['com_name'])
            com_price.append(li_com['com_price'])
            com_stock.append(li_com['com_stock'])

        testData = [com_code, com_name, com_price, com_stock]
        name = '/商品批量导入模板.xlsx'
        filepath = path + name
        dfData = {
            '商品编码': testData[0],
            '商品名称': testData[1],
            '商品价格': testData[2],
            '商品库存': testData[3],
        }
        df = DataFrame(dfData)        # 创建Dataframe
        df.to_excel(filepath, index=False)
        if len(df) != 0:
            messagebox.showinfo('提示', '模板下载成功')
        else:
            messagebox.showinfo('提示', '模板下载失败')

    # 修改商品界面
    def index_modify(self):
        modify = tk.Toplevel
        w = 600
        h = 400
        x = (modify.winfo_screenwidth() -w) /2
        y = (modify.winfo_screenheight() - h )/2
        global com_code
        global com_name
        global com_price
        global com_stock
        com_code = tk.StringVar()
        com_name = tk.StringVar()
        com_price = tk.StringVar()
        com_stock = tk.StringVar()
        modify.geometry('%dx%d+%d+%d' % (w, h, x, y))
        modify.title('修改商品')
        modify.resizable(width=False, height=False)
        modify_frame = Frame(modify, bg='#68B8BE', width=600, height=400)
        modify_frame.pack()
        modify_font_1 = tkFont.Font(family='宋体', size=15, weight=tkFont.BOLD)
        Label(modify_frame, text='添加商品信息', font=modify_font_1, bg='#68B8BE').place(x=200, y=150)
        modify_font_2 = tkFont.Font(family='宋体', size=15)
        modify_font_3 = tkFont.Font(family='宋体', size=9)
        Label(modify_frame, text='商品编号', font=modify_font_2).place(x=235, y=125)
        Label(modify_frame, text='注意：商品编号不可重复', font=modify_font_3).place(x=235, y=125)
        Label(modify_frame, text='商品名称', font=modify_font_2, bg='#68B8BE').place(x=100, y=150)
        Label(modify_frame, text='商品价格', font=modify_font_2, bg='#68B8BE').place(x=100, y=210)
        Label(modify_frame, text='商品库存', font=modify_font_2, bg='#68B8BE').place(x=100, y=270)
        Entry(modify_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_code). \
            place(x=300, y=90, height=30, width=200)
        Entry(modify_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_name). \
            place(x=300, y=150, height=30, width=200)
        Entry(modify_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_price). \
            place(x=300, y=210, height=30, width=200)
        Entry(modify_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=self.com_stock). \
            place(x=300, y=270, height=30, width=200)
        Button(modify_frame, text='立即查询', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command=self.modify_inquire_com).place(x=140, y=325, width=120, height=40)
        Button(modify_frame, text='立即修改', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command=self.modify_inquire_com).place(x=340, y=325, width=120, height=40)

    # 04-1 立即查询商品
    def modify_inquire_com(self):
        sql = 'select * from commodity where com_code=%s'% com_code.get()
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.conn.commit()
        if len(result) ==0:
            com_name.set('未查到相关商品')
            com_price.set('未查到相关商品')
            com_stock.set('未查到相关商品')
        else:
            com = commodity(result[0][0], result[0][1], result[0][2], result[0][3])
            com_name.set(com.get_commodity()['com_name'])
            com_price.set(com.get_commodity()['com_price'])
            com_stock.set(com.get_commodity()['com_stock'])

    # 04-2 修改商品
    def modify_com(self):
        sql = "updata commodity set com_name= '%s', com_price = %s, com_stock =%s where com_code=%s" %(
              com_name.get(), com_price.get(), com_stock.get(), com_code.get())
        result = self.cursor.fetchall()
        self.conn.commit()
        if len(result) ==0:
            messagebox.showinfo('提示', '修改成功')
        else:
            messagebox.showinfo('提示', '修改成失败')

    # 05 删除商品
    def index_delete(self):
        delete = tk.Toplevel
        w = 600
        h = 400
        x = (delete.winfo_screenwidth() -w) /2
        y = (delete.winfo_screenheight() - h )/2
        global com_code
        global com_name
        global com_price
        global com_stock
        com_code = tk.StringVar()
        com_name = tk.StringVar()
        com_price = tk.StringVar()
        com_stock = tk.StringVar()
        com_name = '查询中'
        com_price = '查询中'
        com_stock = '查询中'
        delete.geometry('%dx%d+%d+%d' % (w, h, x, y))
        delete.title('修改商品')
        delete.resizable(width=False, height=False)
        delete_frame = Frame(delete, bg='#68B8BE', width=600, height=400)
        delete_frame.pack()
        delete_font_1 = tkFont.Font(family='宋体', size=15, weight=tkFont.BOLD)
        Label(delete_frame, text='删除商品信息', font=delete_font_1, bg='#68B8BE').place(x=200, y=150)
        delete_font_2 = tkFont.Font(family='宋体', size=15)
        Label(delete_frame, text='商品编号', font=delete_font_2, bg='#68B8BE').place(x=200, y=90)
        Label(delete_frame, text='商品名称', font=delete_font_2, bg='#68B8BE').place(x=200, y=150)
        Label(delete_frame, text='商品价格', font=delete_font_2, bg='#68B8BE').place(x=200, y=210)
        Label(delete_frame, text='商品库存', font=delete_font_2, bg='#68B8BE').place(x=200, y=270)
        delete_com_code= tk.StringVar()

        Entry(delete_frame, highlightthickness=1, font=('宋体', 15), bg='F3F3F4', textvariable=delete_com_code). \
            place(x=300, y=90, height=30, width=200)
        com_name_label = Label(delete_frame, text=com_name, font=delete_font_2, bg='#68B8BE')
        com_name_label.place(x=300, y=150)
        com_price_label = Label(delete_frame, text=com_price, font=delete_font_2, bg='#68B8BE')
        com_price_label.place(x=300, y=150)
        com_stock_label = Label(delete_frame, text=com_stock, font=delete_font_2, bg='#68B8BE')
        com_stock_label.place(x=300, y=150)

        Button(delete_frame, text='立即查询', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command=self.delete_inquire_com).place(x=140, y=325, width=120, height=40)
        Button(delete_frame, text='立即删除', font=('宋体', 15, 'bold'), fg='#000000', bg='#ffffff',
               command=self.delete_com).place(x=340, y=325, width=120, height=40)

    # 05-1 删除查询商品
    def delete_inquire_com(self):
        sql = 'select * from commodity where com_code=%s'% com_code.get()
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        self.conn.commit()
        if len(result) !=0:
            com = commodity(result[0][0], result[0][1], result[0][2], result[0][3])
            com_name_label.config(text=com.get_commodity()['com_name'])
            com_price_label.config(text=com.get_commodity()['com_price'])
            com_stock_label.config(text=com.get_commodity()['com_stock'])
        else:
            com = commodity(result[0][0], result[0][1], result[0][2], result[0][3])
            com_name_label.config(text='未查询到相关商品')
            com_price_label.config(text='未查询到相关商品')
            com_stock_label.config(text='未查询到相关商品')

    # 05-2 删除商品
    def delete_com(self):
        sql = "delete from commodity where com_code=%s" % com_code.get()
        result = self.cursor.fetchall()
        self.conn.commit()
        if len(result) ==0:
            messagebox.showinfo('提示', '删除成功')
        else:
            messagebox.showinfo('提示', '删除失败')

    # 06 关于界面
    def index_about(self):
        about = Tk()
        w = 600
        h = 400
        x = (about.winfo_screenwidth() -w) /2
        y = (about.winfo_screenheight() - h )/2
        about.geometry('%dx%d+%d+%d' % (w, h, x, y))
        about.title('版权信息')
        about.resizable(width=False, height=False)
        about_frame = Frame(about, bg='#68B8BE', width=600, height=400)
        about_frame.pack()
        about_font_1 = tkFont.Font(family='宋体', size=15, weight=tkFont.BOLD)
        Label(about_frame, text='Python-Excel 超市信息管理系统', font=about_font_1, bg='#68B8BE').place(x=200, y=150)
        Label(about_frame, text='Copy by Master wang', font=about_font_1, bg='#68B8BE').place(x=200, y=150)
        Label(about_frame, text='仅限学习和交流使用', font=about_font_1, bg='#68B8BE').place(x=200, y=210)
        about.mainloop()

    # 07 退出
    def index_quit(self):
        self.conn.close()
        self.master.destroy()


