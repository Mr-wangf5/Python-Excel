from openpyxl.styles import Font
from openpyxl import Workbook
import openpyxl
import tkinter.messagebox

class SMS(object):
    # 01 初始化
    def __init__(self):
        self.student_list = []
        self.font = Font(name='Courier New', size=9, italic=True, bold=False)
        self.student_wb = openpyxl.load_workbook('Files/student.xlsx')
        self.grade_wb = openpyxl.load_workbook('Files/Grade.xlsx')

        # 学生信息表
        self.snames = self.student_wb.sheetnames
        self.student_sheet = self.student_wb[self.snames[0]]
        self.student_sheet.font = self.font

        # 学生课程表
        self.gnames = self.grade_wb.sheetnames
        self.grade_sheet = self.grade_wb[self.gnames[0]]
        self.grade_sheet.font = self.font

    # 02 获取两个工作表
    def update_all(self):
        self.student_wb = openpyxl.load_workbook('Files/student.xlsx')    # 更新
        self.grade_wb = openpyxl.load_workbook('Files/Grade.xlsx')
        # 学生信息表
        self.snames = self.student_wb.sheetnames
        self.grade_sheet = self.grade_wb[self.gnames[0]]

    # 添加学生信息
    def add_student(self, sName, sNo, sSex, tele_num):
        self.update_all()
        if sName == '' or sNo == '' or sSex == '' or tele_num == '':
            tkinter.messagebox.showinfo('提示', '请输入完整信息')
        else:
            r = self.student_sheet.max_row
            c = self.student_sheet.max_column
            s = str(self.student_sheet.cell(row=r, column=1).value)
            row = [self.student_sheet.max_row, sNo, sSex, tele_num]
            self.student_sheet.append(row)
            self.student_wb.save('Files/student.xlsx')
            tkinter.messagebox.showinfo('提示', '插入成功')

    # 添加成绩
    def add_grade(self, sNo, sName, sLanguage, sMath, sOLanguage, sPhysics, sChemistry, sHenwu):
        self.update_all()
        if sName == '' or sNo == '' or sLanguage == '' or sMath == '' or sOLanguage == '' or sPhysics == '' \
                or sChemistry == ''or sHenwu == '':
            tkinter.messagebox.showinfo('提示', '请输入完整信息')
        else:
            r = self.student_sheet.max_row
            s = str(self.student_sheet.cell(row=r, column=1).value)
            row1 = [self.student_sheet.max_row, sNo, sName, sLanguage, sMath, sOLanguage, sPhysics, sChemistry, sHenwu]
            self.grade_sheet.append(row1)
            self.student_wb.save('Files/Grade.xlsx')
            tkinter.messagebox.showinfo('提示', '插入成功')

    # 判断学号是否存在
    def judge_student(self, sno):
        self.update_all()
        index1 = -1
        r = self.student_sheet.max_row
        for i in range(2, r+1):
            s = str(self.student_sheet.cell(row=i, column=2).value)
            if s == str(sno):
                tkinter.messagebox.showinfo('提示', '学生表学号存在')
                return i
        tkinter.messagebox.showinfo('提示', '学生表学号不存在')
        return index1

    # 判断在成绩表中
    def judge_grade(self, sno):
        self.update_all()
        index1 = -1
        r = self.grade_sheet.max_row
        for i in range(2, r+1):
            s = str(self.grade_sheet.cell(row=i, column=2).value)
            if s == str(sno):
                tkinter.messagebox.showinfo('提示', '学生表学号存在')
                return i
        tkinter.messagebox.showinfo('提示', '学生表学号不存在')
        return index1
    # 判断两张表
    def judge_All(self, sno):
        flag = False
        self.update_all()
        num1, num2 = -1, -1
        r1 = self.student_sheet.max_row
        r2 = self.grade_sheet.max_row
        for i in range(1, r1):
            s1 = str(self.student_sheet.cell(row=i, column=2).value)
            if s1 == str(sno):
                tkinter.messagebox.showinfo('提示', '学号存在')
                num1 = i
        if num1 ==-1:
            tkinter.messagebox.showinfo('提示', '学生信息表学号不存在')

        for i in range(1, r2):
            s2 = str(self.grade_sheet.cell(row=i, column=2).value)
            if s2 == str(sno):
                tkinter.messagebox.showinfo('提示', '学号存在')
                num2 = i
        if num2 ==-1:
            tkinter.messagebox.showinfo('提示', '学生信息表学号不存在')
        return num1, num2

    # 修改学生信息
    def modify_student(self, index1, sName, sNo, sSex, tele_num):
        self.update_all()
        if index1 != 1:
            self.student_sheet.cell(row=index1, column=2).value = sNo
            self.student_sheet.cell(row=index1, column=3).value = sName
            self.student_sheet.cell(row=index1, column=4).value = sSex
            self.student_sheet.cell(row=index1, column=5).value = tele_num
            self.student_wb.save('Files/Student.xlsx')
            tkinter.messagebox.showinfo('提示', '修改成功')
        else:
            tkinter.messagebox.showinfo('提示', '学号不存在或请先检索')

    # 修改学生成绩
    def modify_grade(self, index1, sNo, sName, sLanguage, sMath, sOLanguage, sPhysics, sChemistry, sHenwu):
        self.update_all()
        if index1 != 1:
            self.student_sheet.cell(row=index1, column=1).value = self.grade_sheet.max_row
            self.student_sheet.cell(row=index1, column=2).value = sNo
            self.student_sheet.cell(row=index1, column=3).value = sName
            self.student_sheet.cell(row=index1, column=4).value = sLanguage
            self.student_sheet.cell(row=index1, column=5).value = sMath
            self.student_sheet.cell(row=index1, column=6).value = sOLanguage
            self.student_sheet.cell(row=index1, column=7).value = sPhysics
            self.student_sheet.cell(row=index1, column=8).value = sChemistry
            self.student_sheet.cell(row=index1, column=9).value = sHenwu
            self.grade_wb.save('Files/Grade.xlsx')
            tkinter.messagebox.showinfo('提示', '修改成功')
        else:
            tkinter.messagebox.showinfo('提示', '学号不存在或请先检索')

    # 10 删除学生信息表
    def delete_student(self, index1):
        self.update_all()
        if index1 != -1:
            self.update_all()
            self.student_sheet.delete_rows(index1)
            self.student_wb.save('Files/Student.xlsx')
            tkinter.messagebox.showinfo('提示', '删除成功')
        else:
            tkinter.messagebox.showinfo('提示', '学号不存在或请先检索')


    # 11 删除学生成绩
    def delete_grade(self, index1):
        self.update_all()
        if index1 != -1:
            self.update_all()
            self.grade_sheet.delete_rows(index1)
            self.grade_wb.save('Files/Student.xlsx')
            tkinter.messagebox.showinfo('提示', '删除成功')
        else:
            tkinter.messagebox.showinfo('提示', '学号不存在或请先检索')

    # 12 查找学生信息
    def search_student(self, sno):
        self.update_all()
        str1 = ' '
        flag = False
        index = -1
        r = self.student_sheet.max_row
        for i in range(3, r+1):
            s = str(self.student_sheet.cell(row=i, column=2).value)
            if s == str(sno):
                flag = True
                index = i
            else:
                pass
        if flag:
            str1='学号'+str(self.student_sheet.cell(row=index, column=2).value)+'姓名'+str(
                self.student_sheet.cell(row=index, column=3).value)+ '性别' + str(
                self.student_sheet.cell(row=index, column=4).value)+'电话'+ str(
                self.student_sheet.cell(row=index, column=5).value)
            tkinter.messagebox.showinfo('学生信息', str1)
        else:
            tkinter.messagebox.showinfo('错误提示', '学号不存在')

    # 12 查找学生成绩
    def search_grade(self, sno):
        self.update_all()
        str1 = ''
        flag = False
        index = -1
        r = self.student_sheet.max_row
        for i in range(3, r+1):
            s = str(self.student_sheet.cell(row=i, column=2).value)
            if s == str(sno):
                flag = True
                index = i
            else:
                pass
        if flag:
            str1 = '学号'+str(self.student_sheet.cell(row=index, column=2).value)+'姓名'+str(
                self.student_sheet.cell(row=index, column=3).value) + '语文' + str(
                self.student_sheet.cell(row=index, column=4).value)+'数学' + str(
                self.student_sheet.cell(row=index, column=5).value)+'英语' + str(
                self.student_sheet.cell(row=index, column=6).value)+'物理' + str(
                self.student_sheet.cell(row=index, column=7).value)+'化学' + str(
                self.student_sheet.cell(row=index, column=8).value)+'生物' + str(
                self.student_sheet.cell(row=index, column=9).value)
            tkinter.messagebox.showinfo('学生成绩', str1)
        else:
            tkinter.messagebox.showinfo('错误提示', '学号不存在')

    # 14 显示所有学生
    def display_students(self):
        self.update_all()
        str1 = ''
        for i in range(1, self.student_sheet.max_row+1):
            for j in range(1, self.student_sheet.max_column+1):
                s = str(self.student_sheet.cell(row=i, column=j).value)
                if s == 'None':
                    continue
                str1 += s + '\t'
            str1 += '\n'
        return str1


    # 15 显示所有学生成绩
    def display_students(self):
        self.update_all()
        str2 = ''
        for i in range(1, self.grade_sheet.max_row+1):
            for j in range(1, self.grade_sheet.max_column+1):
                s = str(self.grade_sheet.cell(row=i, column=j).value)
                if s == 'None':
                    continue
                str2 += s + '\t'
            str2 += '\n'
        return str2

    # 16 创建个人信息表
    def crate_personal_xlsx(self, num1, num2):
        self.update_all()
        if num1 != -1 and num2 != -1:
            wb = Workbook()
            ws = wb.active
            name = self.student_sheet.cell(row=num1, column=3).value
            sno = self.student_sheet.cell(row=num1, column=2).value
            sex = self.student_sheet.cell(row=num1, column=4).value
            tel = self.student_sheet.cell(row=num1, column=5).value
            datalist = []
            datajige= [0, 0, 0, 0, 0]
            sum = 0
            for i in range(4, 10):
                num = self.grade_sheet.cell(row=num2, column=i).value
                datalist.append(num)
                sum = sum + int(num)
                if int(num) < 60:
                    datajige[0] +=1
                if int(num) < 70 and int(num)>=60:
                    datajige[1]+=1
                if int(num) < 80 and int(num)>=70:
                    datajige[2]+=1
                if int(num) < 90 and int(num)>=80:
                    datajige[3]+=1
                if int(num)>=90:
                    datajige[4]+=1
            strZ = ''
            if sum >= 540:
                strZ = '优'
            elif sum < 540 and sum >= 480:
                strZ = '良'
            elif sum < 480 and sum >= 420:
                strZ = '中'
            elif sum < 420 and sum >= 360:
                strZ = '及格'
            else:
                strZ = '不及格'

            average = sum // 6
            jidian = average / 10 -6
            data = [
                ['学生个人信息', '', '', ''],
                ['姓名', name, '学号', sno],
                ['性别', sex, '电话', tel],
                ['语文', datalist[0], '物理', datalist[3]],
                ['数学', datalist[1], '物理', datalist[4]],
                ['外语', datalist[2], '物理', datalist[5]],
                ['学生成绩分析', '', '', ''],
                ['优科数目', datalist[4], '良科数目', datalist[3]],
                ['中科数目', datalist[2], '及格科数目', datalist[1]],
                ['不及格科数目', datalist[0], '', ''],
                ['学生成绩分析', '', '', ''],
                ['总成绩', sum, '平均成绩', average],
                ['绩点', jidian, '总评', strZ],
            ]
            i = 1
            j = 1
            for line in data:
                for s in line:
                    ws.cell(row=i, column=j).value = str(s)
                    j += 1
                i +=1
                j = 1
            wb.save('Files'+str(sno)+str(name)+'.xlsx')
            tkinter.messagebox.showinfo('保存提示', '保存为相对路径下:Files/学号.xlsx')
            tkinter.messagebox.showinfo('提示', '学生个人信息导出成功')
        else:
            tkinter.messagebox.showinfo('提示', '学号不存在或请先检索')